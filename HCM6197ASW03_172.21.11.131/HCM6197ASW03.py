import textfsm as tf
import openpyxl

# this log file doesn't need to be standalized interfacename
def interface_name(name):
    return 

def read_data(fileName, start_marker, end_marker):
    f = open(fileName, "r")
    file_content = f.read()

    start_index = file_content.find(start_marker) + len(start_marker)
    end_index = file_content.find(end_marker, start_index)

    relevant_data = file_content[start_index:end_index].strip()
    return relevant_data

if (__name__ == '__main__'):

    # get log file name and device information
    fileName = "HCM6197ASW03_172.21.11.131_S5328.txt"
    deviceInfos = fileName.split("_")
    deviceName = deviceInfos[0]
    deviceIP = deviceInfos[1]
    deviceModel = deviceInfos[2]

    # open interface template
    with open("HCM6197ASW03_interfaces.template") as interface_tpl:
        fsm = tf.TextFSM(interface_tpl)

    # open description template
    with open("HCM6197ASW03_description.template") as description_tpl:
        fsm2 = tf.TextFSM(description_tpl)
    
    # read lines in log file with interface info
    start_marker = "Interface                   PHY   Protocol InUti OutUti   inErrors  outErrors"
    end_marker = "@@BLOCK--"
    interfaces_data = read_data(fileName, start_marker, end_marker)
    interfaces_results = fsm.ParseText(interfaces_data)

    # read lines in log file with description info
    start_marker = "<HCM6197ASW03>display interface description"
    end_marker = "@@BLOCK--"
    description_data = read_data(fileName, start_marker, end_marker)
    description_results = fsm2.ParseText(description_data)

    # create dictionary by description
    description_dict = {record[0]: record[1] for record in description_results}

    # save record to DataCollection.xlsx
    workbook = openpyxl.load_workbook("..\DataCollection.xlsx")

    sheetName = str(deviceName)
    if (sheetName in workbook.sheetnames) == True:
        workbook.remove(workbook[sheetName]) 
    worksheet = workbook.create_sheet(sheetName)
    title = ["Interface", "Physical", "Protocol", "InUti", "OutUti", 
             "inErrors", "outErrors", "Description"]
    for col, val in enumerate(title, start=1):
        worksheet.cell(row=1, column=col).value = val
    row_num=2
    for result in interfaces_results:
        interface = result[0]
        phy = result[1]
        protocol = result[2]
        in_uti = result[3]
        out_uti = result[4]
        in_err = result[5]
        out_err = result[6]
        description = description_dict.get(interface, "N/A")

        line = [interface, phy, protocol, in_uti, out_uti, in_err, out_err, description]
        for col_num, res in enumerate(line, start=1):
            worksheet.cell(row_num, col_num).value=res
        row_num = row_num + 1
        
    workbook.save("..\DataCollection.xlsx")
    workbook.close()
