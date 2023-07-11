import textfsm as tf
import openpyxl

# import ReadLogFile Module
import os, sys
script_dir = os.path.dirname( __file__ )
mymodule_dir = os.path.join( script_dir, '..')
sys.path.append( mymodule_dir )
import ReadLogFile

# standalize interface name
def interface_name(name):
    return name.replace("gei_", "GigabitEthernet")

if (__name__ == '__main__'):
    # file name and device information
    fileName = "VTU0009ASW01_172.16.32.248_5928E.txt"
    deviceInfos = fileName.split('_')
    deviceName = deviceInfos[0]
    deviceIP = deviceInfos[1]
    deviceModel = deviceInfos[2]

    # Template
    with open(deviceName + ".template") as tpl:
        fsm = tf.TextFSM(tpl)
    
    # Start marker, end marker and read log data
    start_marker = "Interface     AdminStatus  PhyStatus  Protocol  Description"
    end_marker = "@@BLOCK--"
    data = ReadLogFile.read_data(fileName, start_marker, end_marker)
    results = fsm.ParseText(data)

    # save results to new sheet
    workbook = openpyxl.load_workbook("..\DataCollection.xlsx")
    sheetName = str(deviceName)
    if (sheetName in workbook.sheetnames) == True:
        workbook.remove(workbook[sheetName]) 
    worksheet = workbook.create_sheet(sheetName)
    title = ["Interface", "AdminStatus", "PhysicalStatus", "Protocol", "Description"]

    for col, val in enumerate(title, start=1):
        worksheet.cell(row=1, column=col).value = val
    row_num=2
    for result in results:
        interface = interface_name(result[0])
        admin = result[1]
        phy = result[2]
        protocol = result[3]
        des = result[4]
        line = [interface, admin, phy, protocol, des]
        for col_num, res in enumerate(line, start=1):
            worksheet.cell(row_num, col_num).value=res
        row_num = row_num + 1
        
    workbook.save("..\DataCollection.xlsx")
    workbook.close()