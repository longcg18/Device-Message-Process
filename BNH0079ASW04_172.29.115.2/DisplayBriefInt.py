import textfsm as tf
import openpyxl

import os, sys
script_dir = os.path.dirname( __file__ )
mymodule_dir = os.path.join( script_dir, '..')
sys.path.append( mymodule_dir )
import ReadLogFile

def interface_name(any):
    ats = str(any)
    ats = ats.replace("Eth", "Ethernet").replace("GE", "GigabitEthernet")
    ats = ats.replace("TENGE", "tenGigabitEthernet").replace("Loop", "Loopback")
    return ats

if (__name__ == '__main__'):

    fileName = "BNH0079ASW04 _172.29.115.2_S3900.txt"
    deviceInfos = fileName.split("_")
    deviceName = deviceInfos[0]
    deviceIP = deviceInfos[1]
    deviceModel = deviceInfos[2]


    start_marker = "Interface   Link"
    end_marker = "@BLOCK--"
    relevant_data = ReadLogFile.read_data("BNH0079ASW04 _172.29.115.2_S3900.txt", start_marker, end_marker)

    with open("DisplayBriefInt.template") as f:
        interface_template = tf.TextFSM(f)
    results = interface_template.ParseText(relevant_data)
    
    workbook = openpyxl.load_workbook("..\DataCollection.xlsx")
    sheetName = str(deviceName)
    if (sheetName in workbook.sheetnames) == True:
        workbook.remove(workbook[sheetName]) 
    worksheet = workbook.create_sheet(sheetName)    

    title = ['Interface', 'LinkState', 'SwitchPortMode', 'Description']
    for col, val in enumerate(title, start=1):
        worksheet.cell(row=1, column=col).value = val
    row_num=2
    for result in results:
        
        line = []
        line.append(str(interface_name(result[0])))    
        line.append(str(result[1]))        
        line.append(str(result[2]))  
        line.append(str(result[3]))      
        
        for col_num, res in enumerate(line, start=1):
            worksheet.cell(row_num, col_num).value=res
        row_num = row_num + 1
        
    workbook.save("..\DataCollection.xlsx")
    workbook.close()

