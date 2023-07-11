import textfsm as tf
import xlsxwriter
#from ExcelOpener import open_workbook
#import ExcelOpener

import openpyxl
import os, sys
script_dir = os.path.dirname( __file__ )
mymodule_dir = os.path.join( script_dir, '..')
sys.path.append( mymodule_dir )
import ReadLogFile


def interface_name(name):
    ats = str(name)
    return ats.replace("fei_", "FastEthernet").replace("gei_", "GigabitEthernet")

def read_data(fileName):
    f = open(fileName, "r")
    start_marker = "@@BLOCK--"
    end_marker = "@@BLOCK--"

    file_content = f.read()

    start_index = file_content.find(start_marker) + len(start_marker)
    end_index = file_content.find(end_marker, start_index)

    relevant_data = file_content[start_index:end_index].strip()
    return relevant_data

if (__name__ == '__main__'):

    # Log file
    fileName = "DNG5031ASW09_172.20.200.150_3928A.txt"
    deviceInfos = fileName.split("_")
    deviceName = deviceInfos[0]
    deviceIP = deviceInfos[1]
    deviceModel = deviceInfos[2]

    # Template
    with open("DNG5031ASW09.template") as tpl:
        fsm = tf.TextFSM(tpl)
    
    # Read log file and Parse
    start_marker = "@@BLOCK--"
    end_marker = "@@BLOCK--"
    relevant_data = ReadLogFile.read_data(fileName, start_marker, end_marker)
    results = fsm.ParseText(relevant_data)
    
    # Save results to DataCollection.xlsx
    workbook = openpyxl.load_workbook("..\DataCollection.xlsx")

    sheetName = str(deviceName)
    if (sheetName in workbook.sheetnames) == True:
        workbook.remove(workbook[sheetName]) 
    worksheet = workbook.create_sheet(sheetName)
    title = ["Interface", "BandWidth(Mbits)", "AdminState", "PhysicalState", "ProtocolState", "Description"]
    for col, val in enumerate(title, start=1):
        worksheet.cell(row=1, column=col).value = val
    row_num=2
    for result in results:
        
        line = []
        line.append(str(interface_name(result[0])))    
        line.append(str(result[1]))        
        line.append(str(result[2]))  
        line.append(str(result[3]))      
        line.append(str(result[4]))
        line.append(str(result[5]))
        
        for col_num, res in enumerate(line, start=1):
            worksheet.cell(row_num, col_num).value=res
        row_num = row_num + 1
        
    workbook.save("..\DataCollection.xlsx")
    workbook.close()
