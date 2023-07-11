import textfsm as tf
import openpyxl

# Import module ReadLogFile
import os, sys
script_dir = os.path.dirname( __file__ )
mymodule_dir = os.path.join( script_dir, '..')
sys.path.append( mymodule_dir )
import ReadLogFile

# Main 

if (__name__ == '__main__'):
    
    # Get device information
    fileName = "deviceName_deviceIP_deviceModel.txt"
    deviceInfos = fileName.split('_')
    deviceName = deviceInfos[0]
    deviceIP = deviceInfos[1]
    deviceModel = deviceInfos[2]

    # Read Template
    with open("example.template") as template:
        fsm = tf.TextFSM(template)

    # Read data from log
    start_marker = "Interface       Status              Description"
    end_marker = "@@BLOCK--"
    data = ReadLogFile.read_data(fileName, start_marker, end_marker)

    # Parse to results
    results = fsm.ParseText(data)

    # Write results to WorkSheet (xlsx file)
    workbook = openpyxl.load_workbook("..\Your_work_book.xlsx")

    ## Create new sheet
    sheetName = str(deviceName)
    if (sheetName in workbook.sheetnames) == True:
        workbook.remove(workbook[sheetName]) 
    worksheet = workbook.create_sheet(sheetName)

    ## Set title for columns
    title = ["Interface", "Status", "Description"]
    for col, val in enumerate(title, start=1):
        worksheet.cell(row=1, column=col).value = val
    row_num=2
    for result in results:
        interface = result[0]
        status = result[1]
        description = result[2]
        line = [interface, status, description]
        for col_num, res in enumerate(line, start=1):
            worksheet.cell(row_num, col_num).value=res
        row_num = row_num + 1
        
    ## Save data and close
    workbook.save("..\Your_work_book.xlsx")
    workbook.close()
