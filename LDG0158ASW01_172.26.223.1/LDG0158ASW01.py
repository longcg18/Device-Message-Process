import textfsm as tf
import openpyxl

import os, sys
script_dir = os.path.dirname( __file__ )
mymodule_dir = os.path.join( script_dir, '..')
sys.path.append( mymodule_dir )
import ReadLogFile

def interface_name(name):
    return

if (__name__ == '__main__'):

    # get device infos
    fileName = "LDG0158ASW01_172.26.223.1_IS2828F.txt"
    deviceInfos = fileName.split("_")
    deviceName = deviceInfos[0]
    deviceIP = deviceInfos[1]
    deviceModel = deviceInfos[2]

    # open interface template
    with open("LDG0158ASW01_interface.template") as int_tpl:
        int_fsm = tf.TextFSM(int_tpl)

    # open description template
    with open("LDG0158ASW01_description.template") as des_tpl:
        des_fsm = tf.TextFSM(des_tpl)

    # read int data
    start_marker = "Port  Admin   Operate        Speed/Duplex  Flowctrl(R/S) Mac-learning Status    up-sta                   up-sustained"
    end_marker = "@@BLOCK--"
    interface_data = ReadLogFile.read_data(fileName, start_marker, end_marker)
    interface_results = int_fsm.ParseText(interface_data)

    # read description 
    start_marker = "Port    Description"
    end_marker = "@@BLOCK--"
    description_data = ReadLogFile.read_data(fileName, start_marker, end_marker)
    description_results = des_fsm.ParseText(description_data)
    
    # create dictionary with description
    description_dict = {record[0]: record[1] for record in description_results}

    # save record to DataCollection.xlsx
    workbook = openpyxl.load_workbook("..\DataCollection.xlsx")
    sheetName = str(deviceName)
    if (sheetName in workbook.sheetnames) == True:
        workbook.remove(workbook[sheetName]) 
    worksheet = workbook.create_sheet(sheetName)
    title = ["PortNumber", "AdminState", "Operate", "Speed/Duplex", "Status", "Description"]
    for col, val in enumerate(title, start=1):
        worksheet.cell(row=1, column=col).value = val
    row_num=2
    for result in interface_results:
        port_number = result[0]
        admin_state = result[1]
        operate = result[2]
        speed_duplex = result[3]
        status = result[4]
        description = description_dict.get(port_number, "N/A")

        line = [port_number, admin_state, operate, speed_duplex, status, description]
        for col_num, res in enumerate(line, start=1):
            worksheet.cell(row_num, col_num).value=res
        row_num = row_num + 1
        
    workbook.save("..\DataCollection.xlsx")
    workbook.close()
