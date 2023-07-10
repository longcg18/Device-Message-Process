import textfsm as tf
import openpyxl

# for ReadLogFile Module
import os, sys
script_dir = os.path.dirname( __file__ )
mymodule_dir = os.path.join( script_dir, '..')
sys.path.append( mymodule_dir )
import ReadLogFile

# standalize interface name
def interface_name(name):
    ats = str(name)
    return ats.replace("FE", "FastEthernet").replace("G-", "Gigabit-")

if (__name__ == '__main__'):
    
    # Log file
    fileName = "HNI5335ASW03_172.26.124.98_DCS-3950.txt"
    deviceInfos = fileName.split("_")
    deviceName = deviceInfos[0]
    deviceIP = deviceInfos[1]
    deviceModel = deviceInfos[2]

    # Template
    with open("ShowIntEthStatus.template") as tpl:
        fsm = tf.TextFSM(tpl)
    
    # Read file and Parse
    start_marker = "Interface  Link/Protocol  Speed   Duplex  Vlan   Type"
    end_marker = "@BLOCK--"
    data = ReadLogFile.read_data("HNI5335ASW03_172.26.124.98_DCS-3950.txt", start_marker, end_marker)
    results = fsm.ParseText(data)

    # Save result to DataCollection.xlsx
    workbook = openpyxl.load_workbook("..\DataCollection.xlsx")
    sheetName = str(deviceName)
    if (sheetName in workbook.sheetnames) == True:
        workbook.remove(workbook[sheetName]) 
    worksheet = workbook.create_sheet(sheetName)
    title = ["Interface", "Link/Protocol State", "Speed", "Duplex", "Vlan", "AliasName"]
    for col, val in enumerate(title, start=1):
        worksheet.cell(row=1, column=col).value = val
    row_num=2
    for result in results:
        line = []
        line.append(str(interface_name(result[6]) + " " + result[0]))        
        line.append(str(result[1] + "/" + result[2]))        
        line.append(str(result[3]))        
        line.append(str(result[4]))
        line.append(str(result[5]))
        line.append(str(result[7]))
        
        for col_num, res in enumerate(line, start=1):
            worksheet.cell(row_num, col_num).value=res
        row_num = row_num + 1
    
    workbook.save("..\DataCollection.xlsx")
    workbook.close()

