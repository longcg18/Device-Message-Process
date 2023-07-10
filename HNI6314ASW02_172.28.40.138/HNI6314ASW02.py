import textfsm
import openpyxl

# for ReadLogFile Module
import os, sys
script_dir = os.path.dirname( __file__ )
mymodule_dir = os.path.join( script_dir, '..')
sys.path.append( mymodule_dir )
import ReadLogFile

def interface_name(name):
    ats = str(name)
    return ats.replace("gi", "GigabitEthernet")

def export_title(title):
    print()

if (__name__ == '__main__'): 
    # Log file
    fileName = "HNI6314ASW02_172.28.40.138_MyPowerS4220.txt"
    deviceInfos = fileName.split("_")
    deviceName = deviceInfos[0]
    deviceIP = deviceInfos[1]
    deviceModel = deviceInfos[2]

    # Template
    with open("HNI6314ASW02.template") as tpl:
        fsm = textfsm.TextFSM(tpl)

    # Read log file and Parse
    start_marker = "@@BLOCK--"
    end_marker = "@@BLOCK--"
    data = ReadLogFile.read_data(fileName, start_marker, end_marker)
    results = fsm.ParseText(data)

    # save results to DataCollection.xlsx
    workbook = openpyxl.load_workbook("..\DataCollection.xlsx")
    sheetName = str(deviceName)
    if (sheetName in workbook.sheetnames) == True:
        workbook.remove(workbook[sheetName])
    worksheet = workbook.create_sheet(sheetName)

    # title
    title = ['Interface', 'LinkState', 'ActSpeed', 'ActDuplex', 'PVid', 'Description']
    for col, val in enumerate(title, start=1):
        worksheet.cell(row=1, column=col).value = val
    row_num = 2
    for result in results:
        line = []
        line.append(str(interface_name(result[0])))
        line.append(str(result[1]))
        line.append(str(result[2]))
        line.append(str(result[3]))
        line.append(str(result[4]))
        line.append(str(result[5]))

        for col_num, res in enumerate(line, start=1):
            worksheet.cell(row_num, col_num).value=res
        row_num = row_num + 1

    workbook.save("..\DataCollection.xlsx")
    workbook.close()