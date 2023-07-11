import textfsm as tf
import openpyxl
import re

import os, sys
script_dir = os.path.dirname( __file__ )
mymodule_dir = os.path.join( script_dir, '..')
sys.path.append( mymodule_dir )
import ReadLogFile


def interface_name(name):
    ats = name.replace("Eth", "Ethernet")
    ats = ats.replace("GE", "GigabitEthernet")
    return ats

if (__name__ == '__main__'):
    fileName = "NTN0022ASW01_172.20.133.81_S3300.txt"
    deviceInfos = fileName.split('_')
    deviceName = deviceInfos[0]
    deviceIP = deviceInfos[1]
    deviceModel = deviceInfos[2]

    # 
    with open (deviceName + "_interface.template") as int_tpl:
        int_fsm = tf.TextFSM(int_tpl)
    
    with open (deviceName + "_long_description.template") as des_tpl:
        des_fsm = tf.TextFSM(des_tpl)

    with open (deviceName + "_short_description.template") as short_des_tpl:
        short_des_fsm = tf.TextFSM(short_des_tpl)

    start_marker = "Interface                   PHY   Protocol InUti OutUti   inErrors  outErrors"
    end_marker = "@@BLOCK--"
    int_data = ReadLogFile.read_data(fileName, start_marker, end_marker)
    int_results = int_fsm.ParseText(int_data)
   
    start_marker = "Interface                     PHY     Protocol Description"
    end_marker = "@@BLOCK--"
    des_data = ReadLogFile.read_data(fileName, start_marker, end_marker)
    des_results = des_fsm.ParseText(des_data)

    short_des_data = ReadLogFile.read_data(fileName, start_marker, end_marker)
    short_des_results = short_des_fsm.ParseText(short_des_data)

    description_dict = {interface_name(record[0]): record[1] for record in short_des_results}
    long_description_dict = {interface_name(record[0]): (record[1] + record[2]) for record in des_results}
    long_description_int = []
    for record in des_results:
        long_description_int.append(interface_name(record[0]))
    workbook = openpyxl.load_workbook("..\DataCollection.xlsx")
    sheetName = str(deviceName)
    if (sheetName in workbook.sheetnames) == True:
        workbook.remove(workbook[sheetName]) 
    worksheet = workbook.create_sheet(sheetName)
    title = ['Interface', 'Physical', 'Protocol', 'InUti', 'OutUti', 'inErrors', 'outErrors', 'Description']
    for col, val in enumerate(title, start=1):
        worksheet.cell(row=1, column=col).value = val
    row_num=2
    for result in int_results:
        int = result[0].strip()
        phy = result[1]
        protocol = result[2]
        in_uti = result[3]
        out_uti = result[4]
        in_err = result[5]
        out_err = result[6]
        if (int in long_description_int):
            des = long_description_dict.get(int)
        else:
            des = description_dict.get(int, "N/A")
        
        line = [int, phy, protocol, in_uti, out_uti, in_err, out_err, des]
        for col_num, res in enumerate(line, start=1):
            worksheet.cell(row_num, col_num).value=res
        row_num = row_num + 1
        
    workbook.save("..\DataCollection.xlsx")
    workbook.close()
