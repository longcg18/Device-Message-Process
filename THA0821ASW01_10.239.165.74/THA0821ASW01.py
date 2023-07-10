import textfsm as tf
import openpyxl

import os, sys
script_dir = os.path.dirname( __file__ )
mymodule_dir = os.path.join( script_dir, '..')
sys.path.append( mymodule_dir )
import ReadLogFile

if (__name__ == '__main__'):
    fileName = "THA0821ASW01_10.239.165.74_MES3500-24F.txt"
    deviceInfos = fileName.split("_")
    deviceName = deviceInfos[0]
    deviceIP = deviceInfos[1]
    deviceModel = deviceInfos[2]

    with open(deviceName + ".template") as tpl:
        fsm = tf.TextFSM(tpl)
    
    start_marker = "THA0821ASW01# show interfaces config 1-28" 
    end_marker = "@@BLOCK--"
    data = ReadLogFile.read_data(fileName, start_marker, end_marker)

    results = fsm.ParseText(data)

    workbook = openpyxl.load_workbook("..\DataCollection.xlsx")
    sheetName = str(deviceName)
    if (sheetName in workbook.sheetnames) == True:
        workbook.remove(workbook[sheetName]) 
    worksheet = workbook.create_sheet(sheetName)
    title = ["Port", "Active", "Name", "PVID", "FlowControl", "Type", "SpeedDuplex", "Priority"]
    for col, val in enumerate(title, start=1):
        worksheet.cell(row=1, column=col).value = val
    row_num=2
    for result in results:
        port = result[0]
        active = result[1]
        name = result[2]
        pvid = result[3]
        flow_control = result[4]
        type_t = result[5]
        speed_duplex = result[6]
        priority = result[7]

        line = [port, active, name, pvid, flow_control, type_t, speed_duplex, priority]
        for col_num, res in enumerate(line, start=1):
            worksheet.cell(row_num, col_num).value=res
        row_num = row_num + 1
        
    workbook.save("..\DataCollection.xlsx")
    workbook.close()
    for result in results:
        print(result)
    print()