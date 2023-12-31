import textfsm as tf
import openpyxl

import os, sys
script_dir = os.path.dirname( __file__ )
mymodule_dir = os.path.join( script_dir, '..')
sys.path.append( mymodule_dir )
import ReadLogFile

if (__name__ == '__main__'):
    fileName = "NAN0184ASW01_10.239.119.26_S3100-28FC.txt"

    deviceInfos = fileName.split('_')
    deviceName = deviceInfos[0]
    deviceIP = deviceInfos[1]
    deviceModel = deviceInfos[2]

    with open(deviceName + ".template") as tpl:
        fsm = tf.TextFSM(tpl)

    start_marker = "Port    Desc   Link shutdn Speed         Pri PVID Mode TagVlan    UtVlan"
    end_marker = "Total entries: 28 ."
    data = ReadLogFile.read_data(fileName, start_marker, end_marker)

    results = fsm.ParseText(data)

    workbook = openpyxl.load_workbook("..\DataCollection.xlsx")
    sheetName = str(deviceName)
    if (sheetName in workbook.sheetnames) == True:
        workbook.remove(workbook[sheetName]) 
    worksheet = workbook.create_sheet(sheetName)
    title = ["Port", "Description", "LinkState", "ShutDownStatus", "Operate", "Speed", "Pri", "Mode", "TagVlan", "UtVlan"]
    for col, val in enumerate(title, start=1):
        worksheet.cell(row=1, column=col).value = val
    row_num=2
    for result in results:
        port = result[0]
        desc = result[1]
        link_state = result[2]
        shutdn = result[3]
        speed = result[4]
        pri = result[5]
        pvid = result[6]
        mode = result[7]
        tag_vlan = result[8]
        ut_vlan = result[9]

        line = [port, desc, link_state, shutdn, speed, pri, pvid, mode, tag_vlan, ut_vlan]
        for col_num, res in enumerate(line, start=1):
            worksheet.cell(row_num, col_num).value=res
        row_num = row_num + 1
        
    workbook.save("..\DataCollection.xlsx")
    workbook.close()
    